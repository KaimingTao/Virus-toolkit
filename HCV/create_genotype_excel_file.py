#!/usr/bin/env python3
"""Create one Excel file listing accession/genotype rows with GenBank metadata."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

DEFAULT_INPUT = Path("Genotype_Reference.csv")
DEFAULT_GENBANK_DIR = Path("genotype_files")
DEFAULT_OUTPUT = Path("genotype_accessions.xlsx")
SOURCE_QUALIFIER_RE = re.compile(r'^/([^=]+)=("?)(.*)$')


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a single Excel file with accession, genotype, and GenBank metadata."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Input genotype reference table (default: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--genbank-dir",
        type=Path,
        default=DEFAULT_GENBANK_DIR,
        help=f"Directory containing GenBank files (default: {DEFAULT_GENBANK_DIR})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output Excel path (default: {DEFAULT_OUTPUT})",
    )
    return parser.parse_args(argv)


def collect_field(lines: list[str], field: str) -> str:
    value_parts: list[str] = []
    collecting = False
    for line in lines:
        if line.startswith(field):
            value_parts.append(line[len(field) :].strip())
            collecting = True
            continue
        if collecting:
            if line.startswith(" " * 12):
                value_parts.append(line.strip())
                continue
            break
    return " ".join(part for part in value_parts if part).strip()


def extract_source_location(lines: list[str]) -> str:
    for line in lines:
        if line.startswith("     source"):
            return line.split("source", 1)[1].strip()
    return ""


def extract_source_qualifiers(lines: list[str]) -> dict[str, str]:
    qualifiers: dict[str, list[str]] = {}
    in_source = False
    current_key = ""

    for line in lines:
        if line.startswith("     source"):
            in_source = True
            continue
        if not in_source:
            continue
        if line.startswith("     ") and not line.startswith("                     "):
            break

        stripped = line.strip()
        if not stripped:
            continue

        if stripped.startswith("/"):
            match = SOURCE_QUALIFIER_RE.match(stripped)
            if not match:
                continue
            current_key = match.group(1)
            value = match.group(3).strip()
            if match.group(2) == '"' and value.endswith('"'):
                value = value[:-1]
            qualifiers.setdefault(current_key, []).append(value)
            continue

        if current_key:
            continuation = stripped
            if continuation.endswith('"'):
                continuation = continuation[:-1]
            qualifiers[current_key][-1] = f"{qualifiers[current_key][-1]} {continuation}".strip()

    return {key: " | ".join(values) for key, values in qualifiers.items()}


def parse_genbank(genbank_path: Path) -> dict[str, str | int]:
    text = genbank_path.read_text(encoding="utf-8")
    lines = text.splitlines()
    locus = ""
    if lines and lines[0].startswith("LOCUS"):
        parts = lines[0].split()
        if len(parts) > 1:
            locus = parts[1]

    sequence = ""
    if "ORIGIN" in text:
        origin = text.split("ORIGIN", 1)[1].split("//", 1)[0]
        sequence = re.sub(r"[^acgtunACGTUN]", "", origin).upper()

    return {
        "Source Location": extract_source_location(lines),
        "Locus": locus,
        "Version": collect_field(lines, "VERSION"),
        "Definition": collect_field(lines, "DEFINITION"),
        "Source": collect_field(lines, "SOURCE"),
        "Organism": collect_field(lines, "  ORGANISM"),
        "Sequence Length": len(sequence),
        "GenBank File": str(genbank_path),
        "Source Qualifiers": extract_source_qualifiers(lines),
    }


def load_rows(input_path: Path, genbank_dir: Path) -> list[dict[str, str | int]]:
    rows: list[dict[str, str | int]] = []
    with input_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        if not reader.fieldnames or "Accession" not in reader.fieldnames or "Genotype" not in reader.fieldnames:
            raise SystemExit(f"Expected tab-delimited columns 'Accession' and 'Genotype': {input_path}")

        for row in reader:
            accession = (row.get("Accession") or "").strip()
            genotype = (row.get("Genotype") or "").strip()
            if not accession:
                continue
            genbank_path = genbank_dir / f"{accession}.gb"
            genbank_info = parse_genbank(genbank_path) if genbank_path.exists() else {
                "Source Location": "",
                "Locus": "",
                "Version": "",
                "Definition": "",
                "Source": "",
                "Organism": "",
                "Sequence Length": "",
                "GenBank File": "",
                "Source Qualifiers": {},
            }
            rows.append(
                {
                    "Accession": accession,
                    "Genotype": genotype,
                    **genbank_info,
                }
            )
    return rows


def source_qualifier_headers(rows: list[dict[str, str | int]]) -> list[str]:
    keys: set[str] = set()
    for row in rows:
        qualifiers = row.get("Source Qualifiers", {})
        if isinstance(qualifiers, dict):
            keys.update(qualifiers.keys())
    return [f"source:{key}" for key in sorted(keys)]


def build_headers(rows: list[dict[str, str | int]]) -> list[str]:
    headers = [
        "Accession",
        "Genotype",
        "Source Location",
        "Locus",
        "Version",
        "Definition",
        "Source",
        "Organism",
        "Sequence Length",
        "GenBank File",
    ]
    headers.extend(source_qualifier_headers(rows))
    return headers


def row_value(row: dict[str, str | int], header: str) -> str | int:
    if header.startswith("source:"):
        key = header.split(":", 1)[1]
        qualifiers = row.get("Source Qualifiers", {})
        return qualifiers.get(key, "") if isinstance(qualifiers, dict) else ""
    return row[header]


def csv_output_path(excel_output_path: Path) -> Path:
    return excel_output_path.with_suffix(".csv")


def write_csv(rows: list[dict[str, str | int]], output_path: Path) -> None:
    headers = build_headers(rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row_value(row, header) for header in headers])


def write_workbook(rows: list[dict[str, str | int]], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Genotype Accessions"

    headers = build_headers(rows)

    for idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=row_value(row, header))

    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 18
    worksheet.column_dimensions["F"].width = 60
    worksheet.column_dimensions["G"].width = 30
    worksheet.column_dimensions["H"].width = 60
    worksheet.column_dimensions["I"].width = 16
    worksheet.column_dimensions["J"].width = 40

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    if not args.input.exists():
        raise SystemExit(f"Input file not found: {args.input}")
    if not args.genbank_dir.exists():
        raise SystemExit(f"GenBank directory not found: {args.genbank_dir}")

    rows = load_rows(args.input, args.genbank_dir)
    csv_path = csv_output_path(args.output)
    write_csv(rows, csv_path)
    write_workbook(rows, args.output)
    print(f"Wrote {len(rows)} accession/genotype rows to {csv_path} and {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
