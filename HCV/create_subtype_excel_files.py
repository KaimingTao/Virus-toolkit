#!/usr/bin/env python3
"""Create one Excel file listing accession/subtype pairs from the ICTV CSV."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

DEFAULT_CSV = Path("ictv_download/ictv_table1_all.csv")
DEFAULT_GENBANK_DIR = Path("subtype_files")
DEFAULT_OUTPUT = Path("subtype_accessions.xlsx")
ACCESSION_SPLIT_RE = re.compile(r"\s*,\s*")
SOURCE_QUALIFIER_RE = re.compile(r'^/([^=]+)=("?)(.*)$')


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a single Excel file with accession and subtype columns."
    )
    parser.add_argument(
        "--input-csv",
        type=Path,
        default=DEFAULT_CSV,
        help=f"Input ICTV CSV path (default: {DEFAULT_CSV})",
    )
    parser.add_argument(
        "--genbank-dir",
        type=Path,
        default=DEFAULT_GENBANK_DIR,
        help=f"Directory containing GenBank files (default: {DEFAULT_GENBANK_DIR})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output Excel path (default: {DEFAULT_OUTPUT})",
    )
    return parser.parse_args(argv)


def collect_field(lines: list[str], field: str) -> str:
    value_parts: list[str] = []
    collecting = False
    for line in lines:
        if line.startswith(field):
            value_parts.append(line[len(field) :].strip())
            collecting = True
            continue
        if collecting:
            if line.startswith(" " * 12):
                value_parts.append(line.strip())
                continue
            break
    return " ".join(part for part in value_parts if part).strip()


def parse_genbank(genbank_path: Path) -> dict[str, str | int]:
    text = genbank_path.read_text(encoding="utf-8")
    lines = text.splitlines()
    locus = ""
    if lines and lines[0].startswith("LOCUS"):
        parts = lines[0].split()
        if len(parts) > 1:
            locus = parts[1]

    sequence = ""
    if "ORIGIN" in text:
        origin = text.split("ORIGIN", 1)[1].split("//", 1)[0]
        sequence = re.sub(r"[^acgtunACGTUN]", "", origin).upper()

    return {
        "Source Location": extract_source_location(lines),
        "Locus": locus,
        "Version": collect_field(lines, "VERSION"),
        "Definition": collect_field(lines, "DEFINITION"),
        "Source": collect_field(lines, "SOURCE"),
        "Organism": collect_field(lines, "  ORGANISM"),
        "Sequence Length": len(sequence),
        "GenBank File": str(genbank_path),
        "Source Qualifiers": extract_source_qualifiers(lines),
    }


def extract_source_location(lines: list[str]) -> str:
    for line in lines:
        if line.startswith("     source"):
            return line.split("source", 1)[1].strip()
    return ""


def extract_source_qualifiers(lines: list[str]) -> dict[str, str]:
    qualifiers: dict[str, list[str]] = {}
    in_source = False
    current_key = ""

    for line in lines:
        if line.startswith("     source"):
            in_source = True
            continue
        if not in_source:
            continue
        if line.startswith("     ") and not line.startswith("                     "):
            break

        stripped = line.strip()
        if not stripped:
            continue

        if stripped.startswith("/"):
            match = SOURCE_QUALIFIER_RE.match(stripped)
            if not match:
                continue
            current_key = match.group(1)
            value = match.group(3).strip()
            if match.group(2) == '"' and value.endswith('"'):
                value = value[:-1]
            qualifiers.setdefault(current_key, []).append(value)
            continue

        if current_key:
            continuation = stripped
            if continuation.endswith('"'):
                continuation = continuation[:-1]
            qualifiers[current_key][-1] = f"{qualifiers[current_key][-1]} {continuation}".strip()

    return {key: " | ".join(values) for key, values in qualifiers.items()}


def load_rows(input_csv: Path, genbank_dir: Path) -> list[dict[str, str | int]]:
    rows: list[dict[str, str | int]] = []
    with input_csv.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            subtype = row["Subtype"].strip()
            for accession in ACCESSION_SPLIT_RE.split(row["Accession number(s)"]):
                accession = accession.strip()
                if accession:
                    genbank_path = genbank_dir / f"{accession}.gb"
                    genbank_info = parse_genbank(genbank_path) if genbank_path.exists() else {
                        "Source Location": "",
                        "Locus": "",
                        "Version": "",
                        "Definition": "",
                        "Source": "",
                        "Organism": "",
                        "Sequence Length": "",
                        "GenBank File": "",
                        "Source Qualifiers": {},
                    }
                    rows.append(
                        {
                            "Accession": accession,
                            "Subtype": subtype,
                            **genbank_info,
                        }
                    )
    return rows


def source_qualifier_headers(rows: list[dict[str, str | int]]) -> list[str]:
    keys: set[str] = set()
    for row in rows:
        qualifiers = row.get("Source Qualifiers", {})
        if isinstance(qualifiers, dict):
            keys.update(qualifiers.keys())
    return [f"source:{key}" for key in sorted(keys)]


def write_workbook(rows: list[dict[str, str | int]], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Subtype Accessions"

    headers = [
        "Accession",
        "Subtype",
        "Source Location",
        "Locus",
        "Version",
        "Definition",
        "Source",
        "Organism",
        "Sequence Length",
        "GenBank File",
    ]
    qualifier_headers = source_qualifier_headers(rows)
    headers.extend(qualifier_headers)
    for idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            if header.startswith("source:"):
                key = header.split(":", 1)[1]
                qualifiers = row.get("Source Qualifiers", {})
                value = qualifiers.get(key, "") if isinstance(qualifiers, dict) else ""
            else:
                value = row[header]
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 18
    worksheet.column_dimensions["F"].width = 60
    worksheet.column_dimensions["G"].width = 30
    worksheet.column_dimensions["H"].width = 60
    worksheet.column_dimensions["I"].width = 16
    worksheet.column_dimensions["J"].width = 40

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    if not args.input_csv.exists():
        raise SystemExit(f"Input CSV not found: {args.input_csv}")
    if not args.genbank_dir.exists():
        raise SystemExit(f"GenBank directory not found: {args.genbank_dir}")

    rows = load_rows(args.input_csv, args.genbank_dir)
    write_workbook(rows, args.output)
    print(f"Wrote {len(rows)} accession/subtype rows with GenBank metadata to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
